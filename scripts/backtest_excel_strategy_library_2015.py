#!/usr/bin/env python3
"""Deterministic audit backtest for the Excel A-share strategy library.
Raw TDX .day data; close signals, next available bar open executions.  See column P.
"""
import os, glob, struct, math, argparse
from multiprocessing import Pool
from collections import defaultdict
from datetime import datetime
import numpy as np
from openpyxl import load_workbook

START=20150101; COMM=.0003; STAMP=.0005; SLIP=.001; MIN_LIST=250; AMT_MIN=50_000_000
IDS=list(range(1,33))+[38]

def sma(x,n):
    z=np.full(len(x),np.nan); cs=np.cumsum(np.insert(x,0,0.)); z[n-1:]=(cs[n:]-cs[:-n])/n; return z
def ema(x,n):
    z=np.full(len(x),np.nan); a=2/(n+1); z[n-1]=np.mean(x[:n])
    for i in range(n,len(x)): z[i]=a*x[i]+(1-a)*z[i-1]
    return z
def rsi(x,n):
    d=np.diff(x,prepend=x[0]); up=np.maximum(d,0); dn=np.maximum(-d,0); au=ema(up,n); ad=ema(dn,n); return 100-100/(1+au/(ad+1e-12))
def atr(h,l,c,n=14):
    prev=np.r_[c[0],c[:-1]]; return ema(np.maximum(h-l,np.maximum(abs(h-prev),abs(l-prev))),n)
def rollmax(x,n):
    out=np.full(len(x),np.nan)
    for i in range(n-1,len(x)): out[i]=np.max(x[i-n+1:i+1])
    return out
def rollmin(x,n):
    out=np.full(len(x),np.nan)
    for i in range(n-1,len(x)): out[i]=np.min(x[i-n+1:i+1])
    return out
def cross(a,b): return (a>b)&(np.r_[False,a[:-1]<=b[:-1]])
def psar(h,l):
    out=np.full(len(h),np.nan); out[0]=l[0]; bull=True; ep=h[0]; af=.02
    for i in range(1,len(h)):
      q=out[i-1]+af*(ep-out[i-1]); q=min(q,l[i-1],l[i-2] if i>1 else l[i-1]) if bull else max(q,h[i-1],h[i-2] if i>1 else h[i-1])
      if bull and l[i]<q: bull=False;q=ep;ep=l[i];af=.02
      elif not bull and h[i]>q: bull=True;q=ep;ep=h[i];af=.02
      else:
       if bull and h[i]>ep: ep=h[i];af=min(.2,af+.02)
       if not bull and l[i]<ep: ep=l[i];af=min(.2,af+.02)
      out[i]=q
    return out

def indicators(o,h,l,c,v):
    ma={n:sma(c,n) for n in (5,10,20,30,50,60,200,252)}; em={n:ema(c,n) for n in (5,10,20,26)}
    A=atr(h,l,c); mid=ma[20]; sd=np.full(len(c),np.nan)
    for i in range(19,len(c)): sd[i]=np.std(c[i-19:i+1])
    up=mid+2*sd; lo=mid-2*sd; dif=em[20]-em[26]; dea=ema(dif[~np.isnan(dif)],9)
    # reconstitute signal EMA while retaining leading NaNs
    dea2=np.full(len(c),np.nan); idx=np.where(~np.isnan(dif))[0]; dea2[idx]=dea
    plus=np.r_[0.,np.maximum(h[1:]-h[:-1],0)]; minus=np.r_[0.,np.maximum(l[:-1]-l[1:],0)]
    tr=np.maximum(h-np.r_[c[0],c[:-1]],np.maximum(abs(l-np.r_[c[0],c[:-1]]),h-l)); pdi=100*ema(plus,14)/(ema(tr,14)+1e-9); mdi=100*ema(minus,14)/(ema(tr,14)+1e-9); adx=ema(100*np.abs(pdi-mdi)/(pdi+mdi+1e-9),14)
    k=100*(c-rollmin(l,14))/(rollmax(h,14)-rollmin(l,14)+1e-9); d=sma(k,3)
    obv=np.cumsum(np.sign(np.diff(c,prepend=c[0]))*v); mf=((2*c-h-l)/(h-l+1e-9))*v; cmf=np.full(len(c),np.nan)
    for i in range(19,len(c)): cmf[i]=mf[i-19:i+1].sum()/(v[i-19:i+1].sum()+1e-9)
    return ma,em,A,up,lo,dif,dea2,pdi,mdi,adx,k,d,obv,cmf

def signal(s,i,x):
 o,h,l,c,v,ma,em,A,up,lo,dif,dea,pdi,mdi,adx,k,d,obv,cmf=x[:19]; vm=x[20]; amt=x[21]
 if i<252 or not (amt[i]>=AMT_MIN and np.isfinite(A[i])): return False
 prior20h=np.max(h[i-20:i]); prior55h=np.max(h[i-55:i]); prior10l=np.min(l[i-10:i]);
 if s==1:return (c[i]>ma[20][i] and c[i-1]<=ma[20][i-1]) and ma[20][i]>ma[20][i-5] and v[i]>vm[i]
 if s==2:return em[5][i]>em[20][i] and em[5][i-1]<=em[20][i-1] and c[i]>ma[60][i]
 if s==3:return ma[50][i]>ma[200][i] and ma[50][i-1]<=ma[200][i-1] and ma[200][i]>=ma[200][i-20]
 if s==4:return c[i]>ma[200][i] and c[i-1]<=ma[200][i-1] and ma[200][i]>=ma[200][i-20]
 if s==5:return c[i]>ma[200][i] and dif[i]>dea[i] and dif[i-1]<=dea[i-1] and dif[i]>0 and dif[i]-dea[i]>0
 if s==6:return pdi[i]>mdi[i] and pdi[i-1]<=mdi[i-1] and adx[i]>25 and adx[i]>adx[i-1] and c[i]>ma[50][i]
 if s==7:return c[i]>prior20h and c[i]>ma[200][i]
 if s==8:return c[i]>prior55h
 if s==9:return (up[i]-lo[i])/ma[20][i] <=np.nanpercentile((up[max(0,i-119):i+1]-lo[max(0,i-119):i+1])/ma[20][max(0,i-119):i+1],10) and c[i]>up[i] and v[i]>1.5*vm[i]
 if s==10:return ma[20][i]>ma[20][i-5] and c[i]>up[i] and adx[i]>20
 if s==11:return c[i]>em[20][i]+2*A[i] and em[20][i]>em[20][i-5]
 if s==12:return c[i]>ma[200][i] and x[19][i]<c[i] and x[19][i-1]>=c[i-1]
 if s==13: # ichimoku: tenkan9/kijun26; cloud spans current (unshifted proxy)
  ten=(np.max(h[i-8:i+1])+np.min(l[i-8:i+1]))/2; kij=(np.max(h[i-25:i+1])+np.min(l[i-25:i+1]))/2; a=(ten+kij)/2; b=(np.max(h[i-51:i+1])+np.min(l[i-51:i+1]))/2
  return c[i]>max(a,b) and c[i-1]<=max(a,b) and ten>kij and a>b
 if s==14:return c[i]>=.97*np.max(h[i-251:i+1]) and c[i]/c[i-252] >= np.nanpercentile(c[i-252:i+1]/c[i-252],80) and v[i]>1.5*vm[i]
 if s==15:return c[i]>ma[200][i] and x[22][i]>30 and x[22][i-1]<=30 and np.min(x[22][i-5:i])<30
 if s==16:return c[i]>ma[200][i] and k[i]<20 and d[i]<20 and k[i]>d[i] and k[i-1]<=d[i-1]
 if s==17:return c[i]>ma[200][i] and l[i]<lo[i] and c[i]>lo[i]
 if s==18:return c[i]>ma[200][i] and c[i]>em[20][i]-2*A[i] and c[i-1]<em[20][i-1]-2*A[i-1]
 if s==19:return c[i]>ma[200][i] and x[23][i]<5
 if s==20:return c[i]>ma[200][i] and x[23][i]<10 and all(np.diff(x[23][i-3:i+1])<0) and x[23][i-3]<60
 if s==21:return c[i]>ma[200][i] and all(np.diff(c[i-3:i+1])<0) and c[i]<ma[5][i]
 if s==22:return i%21==0 and c[i]/c[i-21] <= np.nanpercentile(c[max(252,i-252):i+1]/c[max(252,i-252):i+1][0]-1,20)
 if s==23:return (np.max(h[i-20:i+1])-np.min(l[i-20:i+1]))/np.min(l[i-20:i+1])<=.12 and c[i]>np.max(h[i-20:i]) and v[i]>1.5*vm[i]
 if s==24:return c[i-10]/c[i-30]-1>.15 and np.min(l[i-10:i+1])>=.85*np.max(h[i-30:i]) and v[i]<.7*vm[i] and ma[20][i]>ma[20][i-5] and c[i]>o[i]
 if s==25:return c[i]/c[i-20]-1>.15 and c[i]<=1.05*c[i-20] and c[i]>o[i] and v[i]<vm[i]
 if s==26:return o[i-5]<l[i-6] and o[i]>h[i-1] and v[i]>1.5*vm[i]
 if s==27:return c[i]>np.max(h[i-20:i]) and .10 <= (np.max(h[i-120:i])-np.min(l[i-120:i]))/np.max(h[i-120:i]) <=.35 and v[i]>1.5*vm[i]
 if s==28:return c[i]>np.max(h[i-20:i]) and abs(np.min(l[i-10:i])-np.min(l[i-60:i-10]))/np.min(l[i-60:i-10])<=.03 and v[i]<vm[i]
 if s==29:return c[i]>np.max(h[i-20:i]) and np.min(l[i-60:i])<.95*np.min(l[i-20:i]) and v[i]>1.5*vm[i]
 if s==30:return c[i]/c[i-10]-1>=.10 and (np.max(h[i-10:i])-np.min(l[i-10:i]))/np.max(h[i-10:i])<.5 and c[i]>np.max(h[i-5:i]) and v[i]>1.5*vm[i]
 if s==31:return (h[i]-l[i])<=np.min(h[i-6:i+1]-l[i-6:i+1]) and c[i]>ma[20][i] and ma[20][i]>ma[20][i-5]
 if s==32:return c[i]>prior20h and (obv[i]>=np.max(obv[i-20:i]) or cmf[i]>.1) and v[i]>1.5*vm[i]
 if s==38:return c[i-1]<=ma[5][i-1] and c[i]>ma[5][i] and c[i]>ma[20][i] and c[i-1]<=ma[20][i-1] and dif[i]>dea[i] and dif[i-1]<=dea[i-1]
 return False

def should_exit(s,j,e,x,entry,peak):
 o,h,l,c,v,ma,em,A,up,lo,dif,dea,pdi,mdi,adx,k,d,obv,cmf=x[:19]; pnl=c[j]/entry-1; stop=entry-2*A[e]
 if pnl<=-.20:return True
 if s==1:return c[j]<em[10][j] or c[j]<peak-3*A[j] or (j>1 and c[j]<ma[20][j] and c[j-1]<ma[20][j-1]) or c[j]<stop
 if s==2:return cross(em[20],em[5])[j] or c[j]<np.min(l[max(e,j-20):j+1]) or c[j]<stop
 if s==3:return cross(ma[200],ma[50])[j] or c[j]<.97*ma[200][j]
 if s==4:return c[j]<ma[200][j] or c[j]<stop
 if s==5:return cross(dea,dif)[j] or dif[j]<dea[j] or c[j]<stop
 if s==6:return cross(mdi,pdi)[j] or c[j]<peak-3*A[j] or c[j]<stop
 if s==7:return c[j]<np.min(l[j-10:j]) or c[j]<stop
 if s==8:return c[j]<np.min(l[j-20:j]) or c[j]<stop
 if s in (9,11,32):return c[j]<ma[20][j] or c[j]<peak-3*A[j] or c[j]<stop
 if s==10:return c[j]<ma[20][j] or (j<=e+2 and c[j]<c[e]) or c[j]<stop
 if s==12:return x[19][j]>c[j] or c[j]<stop
 if s==13:return c[j]<ma[26][j] if 26 in ma else c[j]<ma[20][j]
 if s==14:return c[j]<ma[20][j] or c[j]<.9*peak or c[j]<stop
 if s==15:return x[22][j]>55 or c[j]>=ma[20][j] or c[j]<ma[200][j] or c[j]<stop or j-e>=10
 if s==16:return k[j]>80 or c[j]>=ma[20][j] or c[j]<np.min(l[e:j+1]) or c[j]<stop or j-e>=10
 if s==17:return c[j]>=ma[20][j] or c[j]>=up[j] or (j>1 and c[j]<lo[j] and c[j-1]<lo[j-1]) or pnl<=-2.5*A[e]/entry
 if s==18:return c[j]>=em[20][j] or c[j]<ma[200][j] or pnl<=-3*A[e]/entry
 if s in (19,20):return c[j]>=ma[5][j] or x[23][j]>70 or c[j]<ma[200][j] or pnl<=-3*A[e]/entry or j-e>=10
 if s==21:return c[j]>ma[5][j] or c[j]<ma[200][j] or j-e>=10
 if s==22:return j-e>=21 or pnl<=-.20
 if s==23:return c[j]>=entry+(np.max(h[e-20:e+1])-np.min(l[e-20:e+1])) or c[j]<np.min(l[e-20:e+1]) or pnl<=-A[e]/entry or c[j]<peak-3*A[j]
 if s==24:return c[j]>=np.max(h[e-30:e+1]) or c[j]<np.min(l[e-10:e+1]) or c[j]<ma[30][j] or c[j]<peak-3*A[j]
 if s==25:return c[j]>=np.max(h[e-30:e+1]) or c[j]<c[e-20]
 if s==26:return c[j]>=entry+.5*(c[e-5]-c[e-10]) or c[j]<l[e]
 if s==27:return c[j]>=entry+(np.max(h[e-120:e+1])-np.min(l[e-120:e+1])) or c[j]<np.min(l[e-20:e+1]) or pnl<=-.08
 if s==28:return c[j]>=entry+(np.max(h[e-20:e+1])-np.min(l[e-60:e+1])) or c[j]<np.min(l[e-10:e+1])
 if s==29:return c[j]>=entry+(np.max(h[e-20:e+1])-np.min(l[e-60:e+1])) or c[j]<np.min(l[e-20:e+1])
 if s==30:return c[j]>=entry+(c[e]-c[e-10]) or c[j]<np.min(l[e-15:e+1]) or c[j]<peak-3*A[j]
 if s==31:return pnl>=2*(entry-np.min(l[e-6:e+1]))/entry or c[j]<em[5][j] or c[j]<np.min(l[e-6:e+1])
 if s==38:return c[j]<lo[j] or c[j]<stop or (c[j]<ma[5][j] and cross(dea,dif)[j])
 return False

def read_day(path):
 b=open(path,'rb').read(); n=len(b)//32; a=np.frombuffer(b[:n*32],dtype=[('d','<u4'),('o','<u4'),('h','<u4'),('l','<u4'),('c','<u4'),('a','<f4'),('v','<u4'),('r','<u4')])
 good=(a['d']>=START)&(a['o']>0)&(a['c']>0)
 return a['d'][good],a['o'][good]/100.,a['h'][good]/100.,a['l'][good]/100.,a['c'][good]/100.,a['v'][good].astype(float)
def backtest_stock(s,dates,o,h,l,c,v,x=None):
 if x is None:
  ma,em,A,up,lo,dif,dea,pdi,mdi,adx,k,d,obv,cmf=indicators(o,h,l,c,v); sar=psar(h,l); x=(o,h,l,c,v,ma,em,A,up,lo,dif,dea,pdi,mdi,adx,k,d,obv,cmf,sar,sma(v,20),sma(v*c*100,20),rsi(c,14),rsi(c,2))
 out=[]; i=252
 while i<len(c)-2:
  if signal(s,i,x):
   e=i+1; entry=o[e]*(1+SLIP+COMM); peak=c[e]; j=e+1
   # Exit conditions consume close[j]; hence execution must be at j+1 open.
   exit_j=None
   while j<len(c)-1:
    peak=max(peak,c[j])
    if should_exit(s,j,e,x,entry,peak):
     exit_j=j+1
     break
    j+=1
   if exit_j is not None:
    ex=o[exit_j]*(1-SLIP-COMM-STAMP); out.append((int(dates[e]),int(dates[exit_j]),entry,ex,e,exit_j,dates,c)); i=exit_j+1
   else:
    # Unclosed position at data end is not fabricated into a completed trade.
    i=len(c)-1
  else:i+=1
 return out

def metrics(trades):
 # Equal-weight active-position daily sleeve; cash earns 0, daily weights equal among live names.
 daily=defaultdict(list); closed=[]; holds=[]
 for ed,xd,ep,xp,e,j,dates,c in trades:
  closed.append(xp/ep-1); holds.append(j-e)
  prev=ep
  # Include the exit bar.  The former half-open range(e, j) dropped every
  # next-day exit (j == e + 1), leaving trades in the count but no return
  # observations and incorrectly rendering their strategy metrics as zero.
  for q in range(e,j+1):
   val=xp if q==j else c[q]
   daily[int(dates[q])].append(val/prev-1); prev=val
 if not daily:return [0]*10
 eq=1.; curve=[]
 for day in sorted(daily): eq*=1+float(np.mean(daily[day])); curve.append(eq)
 years=max((datetime.strptime(str(max(daily)),'%Y%m%d')-datetime.strptime(str(min(daily)),'%Y%m%d')).days/365.25,1/365.25)
 cagr=eq**(1/years)-1; arr=np.array(curve); dd=float(np.min(arr/np.maximum.accumulate(arr)-1)); rets=np.array([np.mean(daily[z]) for z in sorted(daily)]); sh=float(np.mean(rets)/(np.std(rets)+1e-12)*math.sqrt(252)); cal=cagr/abs(dd) if dd<0 else 0
 wins=np.array(closed); win=float(np.mean(wins>0)); payoff=float(wins[wins>0].mean()/-wins[wins<=0].mean()) if any(wins>0) and any(wins<=0) else 0
 return [cagr,dd,sh,cal,win,payoff,len(closed),2*len(closed)/years,float(np.mean(holds)),cagr]
def process_path(p):
 try:
  dates,o,h,l,c,v=read_day(p)
  if len(c)<MIN_LIST or len(c)<253: return None
  ma,em,A,up,lo,dif,dea,pdi,mdi,adx,k,d,obv,cmf=indicators(o,h,l,c,v)
  x=(o,h,l,c,v,ma,em,A,up,lo,dif,dea,pdi,mdi,adx,k,d,obv,cmf,psar(h,l),sma(v,20),sma(v*c*100,20),rsi(c,14),rsi(c,2))
  return int(dates[-1]), {s:backtest_stock(s,dates,o,h,l,c,v,x) for s in IDS}
 except Exception as ex:
  return ('ERR',p,str(ex))
def main():
 ap=argparse.ArgumentParser(); ap.add_argument('--workbook',default='/mnt/c/Users/Sky.Lu/OneDrive - Thermo Fisher Scientific/Documents/股票/outputs/019fb5b5-d55d-7401-99b3-811bcaf6cfc0/A股交易策略库_45策略.xlsx'); ap.add_argument('--tdx',default='/home/lufanfeng/tdx_data'); args=ap.parse_args()
 paths=[]
 for market,prefixes in [('sh',('sh6',)),('sz',('sz00','sz30')),('bj',('bj92',))]:
  for p in glob.glob(os.path.join(args.tdx,'vipdoc',market,'lday','*.day')):
   if os.path.basename(p).startswith(prefixes): paths.append(p)
 print('eligible .day files',len(paths),flush=True)
 results={s:[] for s in IDS}; latest=START
 with Pool(processes=min(12,os.cpu_count() or 1)) as pool:
  for n,row in enumerate(pool.imap_unordered(process_path,paths,chunksize=4),1):
   if row is None: continue
   if row[0]=='ERR': print('skip',row[1],row[2]); continue
   latest=max(latest,row[0])
   for s in IDS: results[s].extend(row[1][s])
   if n%100==0: print(n,flush=True)
 wb=load_workbook(args.workbook); ws=wb['回测结果模板']
 for r in range(5,ws.max_row+1):
  s=ws.cell(r,1).value
  if s not in results: continue
  m=metrics(results[s]); ws.cell(r,4).value=datetime.strptime(str(START),'%Y%m%d').date(); ws.cell(r,5).value=datetime.strptime(str(latest),'%Y%m%d').date()
  for col,val in zip(range(6,16),m): ws.cell(r,col).value=float(val)
  ws.cell(r,16).value=f'原始TDX日线、SH/SZ/BJ(6/00/30/92)、>=250日、20日均额>=5000万；收盘信号T+1开盘，成本=买卖佣金0.03%+滑点0.10%，卖印花0.05%。规则为脚本确定性代理（形态按文档窗口）；整仓等价处理分批；逐日等权活跃仓位组合，现金0%，未复权且不含ST名称筛选；{len(results[s])}笔。'
  for col in (6,7,10,11,13,15): ws.cell(r,col).number_format='0.00%'
  for col in (8,9,12,14): ws.cell(r,col).number_format='0.00'
 wb.save(args.workbook)
 print('latest',latest,'trades', {s:len(results[s]) for s in IDS})
if __name__=='__main__': main()
